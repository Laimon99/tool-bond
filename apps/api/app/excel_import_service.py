"""Excel import normalization service for PoC."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
import math
import re
from typing import Any, Dict, List, Tuple
import uuid

import openpyxl

from .schema_validation import validate_request


DATE_FORMATS = ("%Y-%m-%d", "%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%m/%d/%Y")


@dataclass
class ExcelImportOptions:
    request_id: str | None
    usd_budget: float
    spot_usdtry_override: float | None
    include_breakdown: bool
    persist_run: bool
    rounding_decimals: int
    usd_flat_rate: float


def _to_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    return float(m.group(0))


def _find_label_value(ws: Any, label: str, value_col: int = 2, max_rows: int = 300) -> Any:
    target = label.strip().lower()
    for r in range(1, min(ws.max_row, max_rows) + 1):
        left = ws.cell(r, 1).value
        if left is None:
            continue
        left_s = str(left).strip().lower()
        if left_s == target:
            return ws.cell(r, value_col).value
    return None


def _parse_coupon_maturity(text: str) -> Tuple[float | None, date | None]:
    m = re.search(r"(\d+(?:\.\d+)?)\s+(\d{2}-[A-Za-z]{3}-\d{4})", text)
    if not m:
        return None, None
    coupon_pct = float(m.group(1))
    try:
        maturity = datetime.strptime(m.group(2), "%d-%b-%Y").date()
    except ValueError:
        return None, None
    return coupon_pct / 100.0, maturity


def _safe_year_date(year: int, month: int, day: int) -> date:
    # Keep this robust for month-end edge cases.
    while day > 27:
        try:
            return date(year, month, day)
        except ValueError:
            day -= 1
    return date(year, month, day)


def _annual_schedule(issue_date: date, maturity_date: date) -> List[date]:
    schedule: List[date] = []
    y = issue_date.year + 1
    while y <= maturity_date.year:
        d = _safe_year_date(y, maturity_date.month, maturity_date.day)
        if d <= maturity_date:
            schedule.append(d)
        y += 1
    if maturity_date not in schedule:
        schedule.append(maturity_date)
    return sorted(set(schedule))


def _parse_curve_swap(content: bytes, warnings: List[str]) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb["Chart"] if "Chart" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: Dict[str, Any] = {}

    trade_date = _to_date(ws["B2"].value)
    if trade_date:
        out["settlement_candidate"] = trade_date

    interpolation_raw = str(ws["B8"].value).strip().lower() if ws["B8"].value else "linear"
    interpolation = interpolation_raw if interpolation_raw in {"linear", "loglinear"} else "linear"

    pillars: List[Dict[str, Any]] = []
    spot_mid: float | None = None
    for r in range(3, ws.max_row + 1):
        tenor = ws.cell(r, 4).value
        end_date = _to_date(ws.cell(r, 6).value)
        bid = _to_float(ws.cell(r, 7).value)
        ask = _to_float(ws.cell(r, 8).value)
        if end_date is None or bid is None or ask is None:
            continue
        pillars.append({"end_date": end_date.isoformat(), "bid": bid, "ask": ask})
        if tenor is not None and str(tenor).strip().upper() == "SPOT":
            spot_mid = (bid + ask) / 2.0

    if pillars:
        out["fx_forward_curve"] = {"interpolation": interpolation, "pillars": pillars}
    else:
        warnings.append("Curve_swap file parsed but no valid FX pillars detected.")

    if spot_mid is None and pillars:
        spot_mid = (pillars[0]["bid"] + pillars[0]["ask"]) / 2.0
        warnings.append("SPOT tenor not found; spot estimated from first FX pillar.")

    if spot_mid is not None:
        out["spot_usdtry"] = spot_mid

    return out


def _parse_bond_storico(content: bytes, warnings: List[str]) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: Dict[str, Any] = {}

    descriptor = str(ws["A1"].value or ws.title or "")
    coupon, maturity = _parse_coupon_maturity(descriptor)
    if coupon is None or maturity is None:
        sheet_hint = str(ws.title or "")
        c2, m2 = _parse_coupon_maturity(sheet_hint)
        coupon = coupon if coupon is not None else c2
        maturity = maturity if maturity is not None else m2

    if maturity is not None:
        issue = _safe_year_date(maturity.year - 5, maturity.month, maturity.day)
        schedule = _annual_schedule(issue, maturity)
        out["bond"] = {
            "coupon_rate_annual": coupon if coupon is not None else 0.275,
            "coupon_frequency": 1,
            "face_value_per_unit": 10000.0,
            "issue_date": issue.isoformat(),
            "maturity_date": maturity.isoformat(),
            "schedule_dates": [d.isoformat() for d in schedule],
            "day_count": "ACT/365F",
        }
    else:
        warnings.append("Could not infer bond maturity from bond_storico. Using fallback dates.")

    settle_raw = _find_label_value(ws, "Valuation Settle Date")
    settle_date = _to_date(settle_raw)
    if settle_date:
        out["settlement_candidate"] = settle_date

    bid_ask_raw = _find_label_value(ws, "Bid / Ask Price")
    if bid_ask_raw:
        m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*/\s*([0-9]+(?:\.[0-9]+)?)", str(bid_ask_raw))
        if m:
            bid = float(m.group(1))
            ask = float(m.group(2))
            out["price_input"] = {"type": "clean_price", "value": (bid + ask) / 2.0}

    if "price_input" not in out:
        priced_using = _find_label_value(ws, "Priced Using")
        if priced_using:
            ym = re.search(r"yield priced at\s*([0-9]+(?:\.[0-9]+)?)%", str(priced_using), flags=re.IGNORECASE)
            if ym:
                out["price_input"] = {"type": "ytm", "value": float(ym.group(1)) / 100.0}

    return out


def _parse_bond_turco(content: bytes, warnings: List[str]) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb["Table Data"] if "Table Data" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: Dict[str, Any] = {}
    for r in range(3, ws.max_row + 1):
        d = _to_date(ws.cell(r, 1).value)
        y = _to_float(ws.cell(r, 2).value)
        if d is not None and y is not None:
            out["latest_yield"] = y / 100.0
            out["latest_yield_date"] = d
            return out
    warnings.append("Bond_tURCO file parsed but no valid yield rows found.")
    return out


def _build_fallback_bond(settlement: date) -> Dict[str, Any]:
    maturity = _safe_year_date(settlement.year + 2, 3, 6)
    issue = _safe_year_date(maturity.year - 5, maturity.month, maturity.day)
    schedule = _annual_schedule(issue, maturity)
    return {
        "coupon_rate_annual": 0.275,
        "coupon_frequency": 1,
        "face_value_per_unit": 10000.0,
        "issue_date": issue.isoformat(),
        "maturity_date": maturity.isoformat(),
        "schedule_dates": [d.isoformat() for d in schedule],
        "day_count": "ACT/365F",
    }


def _synthesize_df_points(settlement: date, schedule_dates: List[str], flat_rate: float) -> List[Dict[str, Any]]:
    points: List[Dict[str, Any]] = []
    for ds in schedule_dates:
        d = date.fromisoformat(ds)
        if d <= settlement:
            continue
        t = (d - settlement).days / 365.0
        df = math.exp(-flat_rate * t)
        points.append({"date": d.isoformat(), "df": df})
    if not points:
        d = _safe_year_date(settlement.year + 1, settlement.month, settlement.day)
        points.append({"date": d.isoformat(), "df": math.exp(-flat_rate)})
    return points


def normalize_excel_files(file_blobs: List[Tuple[str, bytes]], options: ExcelImportOptions) -> Tuple[Dict[str, Any], int]:
    warnings: List[str] = []
    errors: List[Dict[str, str]] = []

    if not file_blobs:
        return {
            "status": "failed",
            "normalized_request": None,
            "warnings": [],
            "errors": [
                {"code": "NO_FILES", "message": "No excel files provided.", "field": "files"},
            ],
        }, 400

    parsed: Dict[str, Any] = {}
    for file_name, content in file_blobs:
        low = file_name.lower()
        try:
            if "curve_swap" in low:
                parsed.update(_parse_curve_swap(content, warnings))
            elif "bond_storico" in low:
                parsed.update(_parse_bond_storico(content, warnings))
            elif "bond_turco" in low or "bond_turco".replace("_", "") in low.replace("_", ""):
                parsed.update(_parse_bond_turco(content, warnings))
            else:
                warnings.append(f"File '{file_name}' ignored by PoC normalizer.")
        except Exception as exc:
            errors.append(
                {
                    "code": "FILE_PARSE_ERROR",
                    "message": f"Failed parsing '{file_name}': {exc}",
                    "field": file_name,
                }
            )

    settlement = parsed.get("settlement_candidate")
    if not isinstance(settlement, date):
        settlement = date.today()
        warnings.append("Settlement date not detected. Using current date as fallback.")

    bond = parsed.get("bond")
    if not isinstance(bond, dict):
        bond = _build_fallback_bond(settlement)
        warnings.append("Bond structure not fully detected. Using fallback bond settings.")

    price_input = parsed.get("price_input")
    if not isinstance(price_input, dict):
        latest_yield = parsed.get("latest_yield")
        if isinstance(latest_yield, (int, float)):
            price_input = {"type": "ytm", "value": float(latest_yield)}
            warnings.append("Using latest yield from Bond_tURCO as fallback price_input.")
        else:
            price_input = {"type": "clean_price", "value": 90.0}
            warnings.append("Price input not detected. Using fallback clean price 90.0.")

    fx_forward_curve = parsed.get("fx_forward_curve")
    if not isinstance(fx_forward_curve, dict):
        errors.append(
            {
                "code": "MISSING_FX_CURVE",
                "message": "FX forward curve not detected. Provide Curve_swap.xlsx.",
                "field": "fx_forward_curve",
            }
        )
        fx_forward_curve = {"interpolation": "linear", "pillars": []}

    if options.spot_usdtry_override is not None:
        spot_usdtry = float(options.spot_usdtry_override)
    else:
        spot_usdtry = parsed.get("spot_usdtry")
        if not isinstance(spot_usdtry, (int, float)):
            spot_usdtry = 40.0
            warnings.append("Spot USDTRY not detected. Using fallback value 40.0.")
    spot_usdtry = float(spot_usdtry)

    df_points = _synthesize_df_points(
        settlement=settlement,
        schedule_dates=bond["schedule_dates"],
        flat_rate=float(options.usd_flat_rate),
    )
    warnings.append("USD discount curve synthesized from flat rate for PoC normalization.")

    request_id = options.request_id or f"req_import_{uuid.uuid4().hex[:10]}"
    normalized_request: Dict[str, Any] = {
        "request_id": request_id,
        "input_mode": "excel_import",
        "valuation": {
            "settlement_date": settlement.isoformat(),
            "usd_budget": float(options.usd_budget),
            "spot_usdtry": spot_usdtry,
            "bond": bond,
            "usd_discount_curve": {"df_points": df_points},
            "fx_forward_curve": fx_forward_curve,
            "price_input": price_input,
            "options": {
                "include_breakdown": bool(options.include_breakdown),
                "persist_run": bool(options.persist_run),
                "rounding_decimals": int(options.rounding_decimals),
                "fx_rate_side": "ask",
            },
        },
    }

    schema_errors = validate_request(normalized_request)
    if schema_errors:
        errors.extend(schema_errors)

    if errors:
        return {
            "status": "failed",
            "normalized_request": normalized_request,
            "warnings": warnings,
            "errors": errors,
        }, 422

    return {
        "status": "success",
        "normalized_request": normalized_request,
        "warnings": warnings,
        "errors": [],
    }, 200
